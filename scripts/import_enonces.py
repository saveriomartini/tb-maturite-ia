"""Reinjecte les enonces rediges depuis le classeur vers le modele JSON.

Usage : python scripts/import_enonces.py
Entree : redaction/enonces_a_rediger.xlsx
Sortie : src/data/modele.json mis a jour (a relire dans git diff avant commit)

Un enonce dont texte_fr est non vide passe au statut 'redige'.
"""
import json
from pathlib import Path

from openpyxl import load_workbook

RACINE = Path(__file__).resolve().parent.parent
MODELE = RACINE / "src" / "data" / "modele.json"
ENTREE = RACINE / "redaction" / "enonces_a_rediger.xlsx"

modele = json.loads(MODELE.read_text(encoding="utf-8"))
par_id = {e["id"]: e for e in modele["enonces"]}

ws = load_workbook(ENTREE).active
colonnes = {c.value: i for i, c in enumerate(ws[1])}
maj = 0
for ligne in ws.iter_rows(min_row=2, values_only=True):
    ident = ligne[colonnes["id_enonce"]]
    texte = ligne[colonnes["texte_fr"]]
    if ident in par_id and texte and str(texte).strip():
        cible = par_id[ident]
        if cible["texte_fr"] != str(texte).strip():
            cible["texte_fr"] = str(texte).strip()
            maj += 1
        if cible["statut_redaction"] != "hors_perimetre_tb":
            cible["statut_redaction"] = "redige"

MODELE.write_text(json.dumps(modele, ensure_ascii=False, indent=2), encoding="utf-8")
redges = sum(1 for e in modele["enonces"] if e["statut_redaction"] == "redige")
total = sum(1 for e in modele["enonces"] if e["statut_redaction"] != "hors_perimetre_tb")
print(f"{maj} enonce(s) mis a jour. Avancement : {redges}/{total} rediges.")
