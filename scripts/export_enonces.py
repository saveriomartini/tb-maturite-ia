"""Exporte les emplacements d'enonces du modele vers un classeur de redaction.

Usage : python scripts/export_enonces.py
Sortie : redaction/enonces_a_rediger.xlsx (une ligne par enonce)

Le JSON reste la source de verite versionnee ; le classeur n'est qu'une
surface de saisie. Ne jamais rediger directement dans le JSON.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

RACINE = Path(__file__).resolve().parent.parent
MODELE = RACINE / "src" / "data" / "modele.json"
SORTIE = RACINE / "redaction" / "enonces_a_rediger.xlsx"

modele = json.loads(MODELE.read_text(encoding="utf-8"))
cas = {ca["id"]: (d, ca) for d in modele["dimensions"] for ca in d["capability_areas"]}
rubrique = {(r["indicateur"], r["niveau"]): r["enonce_generique_fr"] for r in modele["rubrique_niveau"]}
niveaux = {n["rang"]: n["libelle_fr"] for n in modele["echelle"]["niveaux"]}

wb = Workbook()
ws = wb.active
ws.title = "enonces"
entetes = ["id_enonce", "socle", "dimension", "capability_area", "ref_source", "niveau",
           "libelle_niveau", "accountability", "planning", "resourcing",
           "texte_fr", "statut_redaction"]
ws.append(entetes)
for c in ws[1]:
    c.font = Font(bold=True)

for e in modele["enonces"]:
    d, ca = cas[e["capability_area"]]
    ws.append([
        e["id"], e["socle"], d["nom_fr"], ca["nom_fr"], ca["ref_source"], e["niveau"],
        niveaux[e["niveau"]],
        rubrique.get(("Accountability", e["niveau"]), ""),
        rubrique.get(("Planning", e["niveau"]), ""),
        rubrique.get(("Resourcing", e["niveau"]), ""),
        e["texte_fr"], e["statut_redaction"],
    ])

for col, larg in zip("ABCDEFGHIJKL", (14, 8, 30, 34, 22, 8, 16, 46, 46, 46, 70, 20)):
    ws.column_dimensions[col].width = larg
for ligne in ws.iter_rows(min_row=2):
    for cellule in ligne:
        cellule.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"

SORTIE.parent.mkdir(exist_ok=True)
wb.save(SORTIE)
print(f"{len(modele['enonces'])} enonces exportes vers {SORTIE.relative_to(RACINE)}")
